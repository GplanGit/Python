#importa a biblioteca e instancia a planilha
from multiprocessing.sharedctypes import Value
from sqlite3 import Row
from tkinter.tix import COLUMN, ROW
from openpyxl import Workbook
wb=Workbook()

#instancia uma aba atica dentro da planilha
ws=wb.active

# Cria várias abas a planilha
ws1=wb.create_sheet("Mysheet")#insere no final das abas(default)
ws2=wb.create_sheet("Mysheet",0)#insere na primeira posição
ws3=wb.create_sheet("Mysheet",-1)# insere na penultima posição

#Mundando nome das planilhas
ws2.title="Mysheet_01"
ws3.title="Mysheet_02"

#Mudando cor das abas
ws.sheet_properties.tabcolor="1072BA"

#posicionando pelo nome da aba
ws4=wb["Mysheet_01"]

#retorna nomes das abas
print(wb.sheetnames)

#retorna o nome das abas
for sheet in wb:
    print(sheet.title)

#Criar uma cópia de aba
source=wb.active
target=wb.copy_worksheet(source)
print(wb.sheetnames)

#acessando uma célula
c=ws['a4']
ws['a4']=4
d=ws.cell(row=4,column=2,value=10)
print(d.value)

#acessando varias células
