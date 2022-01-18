from openpyxl import Workbook
wb=Workbook()

#pegue a planilha ativa
ws=wb.active

#Os dados podem ser atribuídos diretamente às células
ws['a1']=42

#As linhas também podem ser anexadas
ws.append([1,2,3])

#Os tipos Python serão convertidos automaticamente
import datetime
ws['a2']=datetime.datetime.now()

#Salva o arquivo
wb.save("sample.xlsx")